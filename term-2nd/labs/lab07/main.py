import pandas as pd
from openpyxl import load_workbook

        
def modifyStudentCode():
    wb = load_workbook('Data.xlsx')
    ws = wb['Sheet1']
    ws['A1'] = "MSHS"
    wb.save('Data.xlsx')
    
def readFile(filePath):
    df = pd.read_excel(filePath)
    return df

def printNFirstLines(df):
    n = int(input("Input the amount of lines to print (from the top): "))
    print("The first {amounts} lines are: ".format(amounts = n))
    print(df.head(n))

def printColsLength(df):
    print("The number of columns is: ", df.shape[0])

def printRowsLength(df):
    print("The number of columns is: ", df.shape[1])

def printSelectedColumn(df):
    target = input("Input the target column (name): ")
    print(df[target])

def printSelectedRow(df):
    target = int(input("Input the target row (pos): "))
    print(df.loc[[target - 1]])

def insertCustomRow(df, insertedData):
    pos = int(input("Input the position of the new row: "))
    df.loc[pos - 1.5] = insertedData
    df = df.sort_index().reset_index(drop = True)
    print(df.loc[pos - 2:pos + 2])
 
def getHighScoresBySubjects(df):
    subject = input("Input the subject you want to get high scores list: ")
    if subject not in df:
        print("The subject you are looking for is not in the data.")
        return
    df = df.sort_values(by = subject)
    print(df.tail(5))

def createMeanCol(df):
    df_score = df.drop(["MSHS"], axis = 1)
    df["Mean"] = df_score.mean(axis = 1)

def createRankCol(df):
    df["Rank"] = df["Mean"].rank(method = "min")
    df = df.sort_values(by = ['Rank'])

def calculateType(row):
    scores = row[1: 12]

    if row["Mean"] >= 8:
        if (scores >= 8).all():
            return "Good"
    if row["Mean"] >= 6.5:
        if (scores >= 6.5).all():
            return "Decent"
    if row["Mean"] >= 5:
        if (scores >= 5).all():
            return "Mid"
    return "Weak"
       
def createType(df):
    df["Type"] = ""
    for index, row in df.iterrows():
        df.at[index, "Type"] = calculateType(row)
    print(df)

def outputToExcel(df):
    df.to_excel("ouput.xlsx")
